# intel-monitor/backend/services/importer.py
"""Excel/CSV 批量导入服务：网站目标 + 社交账号目标。

- 模板生成：xlsx（openpyxl），列名固定
- 文件解析：xlsx（openpyxl）/ xls（xlrd）/ csv（csv 模块）
- 列名校验：列名不匹配时返回明确错误提示
- 行级校验：逐行校验必填/格式，返回成功数 + 失败明细
- 去重：按 URL 跳过已存在目标
"""
import csv
import io
import logging
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)

# ── 列名定义（模板固定列名） ─────────────────────────────
WEBSITE_COLUMNS = ["网站名称", "网站URL"]
TARGET_COLUMNS = ["平台", "账号名称", "账号URL"]

# 支持的文件扩展名
SUPPORTED_EXTS = {".xlsx", ".xls", ".csv"}

# 社交账号合法平台
PLATFORMS = ("x", "youtube", "xiaohongshu", "douyin", "weibo", "toutiao", "108community")
PLATFORM_ALIASES = {
    "x": "x", "twitter": "x", "推特": "x",
    "youtube": "youtube", "油管": "youtube", "youtube.com": "youtube",
    "xiaohongshu": "xiaohongshu", "小红书": "xiaohongshu", "rednote": "xiaohongshu",
    "douyin": "douyin", "抖音": "douyin",
    "weibo": "weibo", "微博": "weibo", "sina": "weibo",
    "toutiao": "toutiao", "头条": "toutiao", "今日头条": "toutiao",
    "108community": "108community", "108": "108community", "108社区": "108community", "天台社区": "108community",
}

_URL_SCHEME_RE = re.compile(r"^https?://", re.IGNORECASE)


# ── 模板生成 ─────────────────────────────────────────────

def make_website_template() -> bytes:
    """生成网站导入模板 xlsx（含示例行与批注）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "网站列表"
    ws.append(WEBSITE_COLUMNS)
    ws.append(["示例网站", "https://example.com/"])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_target_template() -> bytes:
    """生成社交账号导入模板 xlsx（含示例行与批注）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "社交账号列表"
    ws.append(TARGET_COLUMNS)
    ws.append(["x", "示例账号", "https://x.com/example"])
    ws.append(["weibo", "示例微博", "https://weibo.com/u/1234567890"])
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 50
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 文件解析 ─────────────────────────────────────────────

def _read_rows(filename: str, data: bytes) -> list[list[str]]:
    """按扩展名解析文件为二维字符串列表（含表头行）。"""
    ext = Path(filename).suffix.lower()
    if ext not in SUPPORTED_EXTS:
        raise ValueError(f"不支持的文件格式「{ext}」，请使用 xlsx / xls / csv")

    if ext == ".csv":
        # 尝试常见编码
        for enc in ("utf-8-sig", "utf-8", "gbk"):
            try:
                text = data.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError("CSV 文件编码无法识别（支持 UTF-8 / GBK）")
        reader = csv.reader(io.StringIO(text))
        rows = [[c.strip() for c in row] for row in reader if any(c.strip() for c in row)]
        if not rows:
            raise ValueError("CSV 文件为空")
        return rows

    if ext == ".xlsx":
        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            if row and any(v is not None and str(v).strip() for v in row):
                rows.append([str(v).strip() if v is not None else "" for v in row])
        if not rows:
            raise ValueError("Excel 文件为空")
        return rows

    # .xls (老格式)
    try:
        import xlrd
    except ImportError:
        raise ValueError("不支持 .xls 文件：缺少 xlrd 库，请转换为 .xlsx 后重试")
    book = xlrd.open_workbook(file_contents=data)
    sheet = book.sheet_by_index(0)
    rows = []
    for r in range(sheet.nrows):
        vals = [str(sheet.cell_value(r, c)).strip() for c in range(sheet.ncols)]
        if any(vals):
            rows.append(vals)
    if not rows:
        raise ValueError("Excel 文件为空")
    return rows


def _validate_headers(rows: list[list[str]], expected: list[str]) -> None:
    """校验表头列名；不匹配时给出明确提示（指出哪个列名不对）。"""
    headers = rows[0]
    if len(headers) < len(expected):
        raise ValueError(
            f"列数不足：需要 {len(expected)} 列（{'、'.join(expected)}），当前只有 {len(headers)} 列"
        )
    wrong = []
    for i, exp in enumerate(expected):
        actual = headers[i].strip() if i < len(headers) else ""
        if actual != exp:
            wrong.append(f"第{i + 1}列应为「{exp}」但为「{actual}」")
    if wrong:
        raise ValueError("表头列名不匹配：" + "；".join(wrong) + "。请下载模板按固定列名整理后重新上传。")


# ── 通用导入执行 ─────────────────────────────────────────

class ImportResult:
    def __init__(self):
        self.total = 0          # 数据行总数（不含表头）
        self.created = 0        # 成功创建
        self.skipped_dup = 0    # 重复跳过
        self.failed = 0         # 校验失败
        self.errors: list[str] = []  # 失败明细（行号+原因）
        self.created_names: list[str] = []

    def to_dict(self) -> dict:
        return {
            "total": self.total,
            "created": self.created,
            "skipped_dup": self.skipped_dup,
            "failed": self.failed,
            "errors": self.errors[:100],
            "created_names": self.created_names[:50],
        }


def import_websites(filename: str, data: bytes, existing_urls: set[str]) -> tuple[ImportResult, list[dict]]:
    """解析网站导入文件，返回 (结果统计, 待插入条目列表)。"""
    rows = _read_rows(filename, data)
    _validate_headers(rows, WEBSITE_COLUMNS)

    result = ImportResult()
    seen_in_file: set[str] = set()
    items: list[dict] = []

    for idx, row in enumerate(rows[1:], start=2):
        result.total += 1
        name = row[0].strip() if len(row) > 0 else ""
        url = row[1].strip() if len(row) > 1 else ""
        errs = []

        if not name:
            errs.append("网站名称为空")
        if not url:
            errs.append("网站URL为空")
        elif not _URL_SCHEME_RE.match(url):
            url = "https://" + url
        if url and len(url) > 500:
            errs.append("网站URL超长(>500)")

        if errs:
            result.failed += 1
            result.errors.append(f"第{idx}行: {'；'.join(errs)}")
            continue

        key = url.lower().rstrip("/")
        if key in existing_urls or key in seen_in_file:
            result.skipped_dup += 1
            continue

        seen_in_file.add(key)
        result.created += 1
        result.created_names.append(name)

        items.append({
            "name": name, "url": url, "css_selector": None,
            "monitor_interval_minutes": 1440, "monitor_hour": 9, "monitor_minute": 0,
            "is_active": True,
        })

    return result, items


def import_targets(filename: str, data: bytes, existing_urls: set[str]) -> tuple[ImportResult, list[dict]]:
    """解析社交账号导入文件，返回 (结果统计, 待插入条目列表)。"""
    rows = _read_rows(filename, data)
    _validate_headers(rows, TARGET_COLUMNS)

    result = ImportResult()
    seen_in_file: set[str] = set()
    items: list[dict] = []

    for idx, row in enumerate(rows[1:], start=2):
        result.total += 1
        platform_raw = row[0].strip() if len(row) > 0 else ""
        account_name = row[1].strip() if len(row) > 1 else ""
        account_url = row[2].strip() if len(row) > 2 else ""
        errs = []

        platform = PLATFORM_ALIASES.get(platform_raw.lower())
        if not platform:
            errs.append(f"平台「{platform_raw or '空'}」不支持，可选：{'/'.join(PLATFORMS)}")
        if not account_name:
            errs.append("账号名称为空")
        if not account_url:
            errs.append("账号URL为空")
        elif not _URL_SCHEME_RE.match(account_url):
            account_url = "https://" + account_url
        if account_url and len(account_url) > 500:
            errs.append("账号URL超长(>500)")

        if errs:
            result.failed += 1
            result.errors.append(f"第{idx}行: {'；'.join(errs)}")
            continue

        key = account_url.lower().rstrip("/")
        if key in existing_urls or key in seen_in_file:
            result.skipped_dup += 1
            continue

        seen_in_file.add(key)
        result.created += 1
        result.created_names.append(account_name)

        items.append({
            "platform": platform, "account_name": account_name, "account_url": account_url,
            "importance": None, "monitor_interval_minutes": 1440, "monitor_hour": 9,
            "monitor_minute": 0, "post_limit": 10, "post_time_range_days": 0,
            "is_active": True,
        })

    return result, items
